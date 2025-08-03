import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import schedule
import time
cookies = {
    '_xsrf': 'PmOYmAmC2fVXnL3sM0Y9S2r2S92uWtLe',
    '_zap': 'fb962adc-65dc-45f2-bc2b-9f69f5b3bd86',
    'd_c0': 'ACBSZ13fwBmPTkjvUsKjFHEduePKCTKvXZ8=|1735220185',
    'captcha_session_v2': '2|1:0|10:1752830747|18:captcha_session_v2|88:TWx6c2tSTFdKbVhNczhtdGEyR1VadmlRODJSUmZGeVh0RkFkS3d1eDJpVm1MQW54c1hwT0FlUGNyS0RHYWtEKw==|72ec5d9e5ea994083239e6cfc549c82c8882b5346d99206954a7bf362b6359bb',
    '__snaker__id': 'ZQ6j3a6z9I7WM3wn',
    'gdxidpyhxdE': 'en8pIsnncUoH9LBZ1W8Xznk9bvCJQS8vJN6y3DKh5Av3b10kdBth7czlEG%2BwTSJO%5CfvGRKG8UGceDZ5qC0Ab3mRzZ59A1BRcyWrg8c2IXXG4d5D4PMeIwsJ6hqagOessZWTz%5C7NCOfIo9%5CQkKJ7zhYcc%2FVW1yf7TctBa1GhLl4KK0PW3%3A1752831649962',
    '__zse_ck': '004_2/Y/nkWQG4VgWj7p=rczc8bJTtd6VPPiqBq=GQVHB/VMtwwNT7YR/Olxs=gnYJpmzZqHLmShD8HoMhnHvlKMGtcp8J0Bs0Eu5Zny4dvsMRhC9rTNDA631Kc1vSzmhtrz-Dg84K76JZK15AmMptDe0AaZiwh13rzXM9LNr82N4jMfld5cfN/ghGFJUA0hVxqo3eHB8EPojIafj7HY+GcRODy+BZaBRNzCFasv39X3+yctm/RWcAFF1vpUmS0X1ZRXjRK+w8gC8W8jrXbtpI8tIg/pZw0jhuIHEpNG8XXAatv4=',
    'z_c0': '2|1:0|10:1752831125|4:z_c0|92:Mi4xZF9jckd3QUFBQUFBSUZKblhkX0FHU1lBQUFCZ0FsVk5rMkZuYVFDTlhJTWFpejRVV21rZUxvbFg1UGdVUzF0UXZ3|86d74e0522b0f6da6a5961fabcdd43c0d2d6d44cd016a4c04dca7526ab4d2029',
    'Hm_lvt_98beee57fd2ef70ccdd5ca52b9740c49': '1752562705,1752830746,1752831978,1752834129',
    'HMACCOUNT': '6690A0DA60280782',
    'q_c1': 'db64099d329349adb445598247e443d1|1752834350000|1752834350000',
    'SESSIONID': 'IDjlWFNb1UNTHiZeHIbAtq3HPpgsaBnLj5Q8MbD2yYM',
    'JOID': 'V10TAEg3UhWRlZeEdjGkQXdAOJBtfhV6_cTRyREAB1jA3fjeEBt1av-TlIJz1Mt5_TcWvGi445H9zHGb0fzgIQI=',
    'osd': 'UVERAE8xXheRkpGIdDGjR3tCOJdrchd6-sLdyxEHAVTC3f_YHBl1bfmfloJ00sd7_TAQsGq45JfxznGc1_DiIQU=',
    'Hm_lpvt_98beee57fd2ef70ccdd5ca52b9740c49': '1752835725',
    'BEC': '6bca8f185b99e85d761c7a0d8d692864',
    'tst': 'h',
}

headers = {
    'accept': '*/*',
    'accept-language': 'zh-CN,zh;q=0.9',
    'priority': 'u=1, i',
    'referer': 'https://www.zhihu.com/hot',
    'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    'x-api-version': '3.0.76',
    'x-requested-with': 'fetch',
    'x-zse-93': '101_3_3.0',
    'x-zse-96': '2.0_GPl7HPGTQBghO2knAR7fG2jt4kt18dFqOcsRYahSbm28M/XMJ3jNVe4TfAcGgF3h',
    'x-zst-81': '3_2.0aR_sn77yn6O92wOB8hPZnQr0EMYxc4f18wNBUgpTQ6nxERFZf0Y0-4Lm-h3_tufIwJS8gcxTgJS_AuPZNcXCTwxI78YxEM20s4PGDwN8gGcYAupMWufIeQuK7AFpS6O1vukyQ_R0rRnsyukMGvxBEqeCiRnxEL2ZZrxmDucmqhPXnXFMTAoTF6RhRuLPF7HKJwg8QDoBFqV8LugCZGSqNJX9-GL1iwe95g39TGe_-wc_VhHqBwcfJhSTvLcKhgx9cw28p9HK6Bx8c_YMJqVLuboVo_OOQRFfwrNBJrH8rQx1TgcV0q2_LrV_TCHLb9xYBcnqHhe8k9x8zqXGBTg1pbxBhGHOWhCBPUwK64O_fCXsCqfzFG3fDugBrTSYEDVykCtqJAc9SqCfjckqKcX0jUomA9cVSwNKxhL9UueVk0X_rXL_hUVKfbS8sUFYiBxMiupLtbXOfJSYX9O9gDwm8vX1CDVYWhHL1XYC',
    # 'cookie': '_xsrf=PmOYmAmC2fVXnL3sM0Y9S2r2S92uWtLe; _zap=fb962adc-65dc-45f2-bc2b-9f69f5b3bd86; d_c0=ACBSZ13fwBmPTkjvUsKjFHEduePKCTKvXZ8=|1735220185; captcha_session_v2=2|1:0|10:1752830747|18:captcha_session_v2|88:TWx6c2tSTFdKbVhNczhtdGEyR1VadmlRODJSUmZGeVh0RkFkS3d1eDJpVm1MQW54c1hwT0FlUGNyS0RHYWtEKw==|72ec5d9e5ea994083239e6cfc549c82c8882b5346d99206954a7bf362b6359bb; __snaker__id=ZQ6j3a6z9I7WM3wn; gdxidpyhxdE=en8pIsnncUoH9LBZ1W8Xznk9bvCJQS8vJN6y3DKh5Av3b10kdBth7czlEG%2BwTSJO%5CfvGRKG8UGceDZ5qC0Ab3mRzZ59A1BRcyWrg8c2IXXG4d5D4PMeIwsJ6hqagOessZWTz%5C7NCOfIo9%5CQkKJ7zhYcc%2FVW1yf7TctBa1GhLl4KK0PW3%3A1752831649962; __zse_ck=004_2/Y/nkWQG4VgWj7p=rczc8bJTtd6VPPiqBq=GQVHB/VMtwwNT7YR/Olxs=gnYJpmzZqHLmShD8HoMhnHvlKMGtcp8J0Bs0Eu5Zny4dvsMRhC9rTNDA631Kc1vSzmhtrz-Dg84K76JZK15AmMptDe0AaZiwh13rzXM9LNr82N4jMfld5cfN/ghGFJUA0hVxqo3eHB8EPojIafj7HY+GcRODy+BZaBRNzCFasv39X3+yctm/RWcAFF1vpUmS0X1ZRXjRK+w8gC8W8jrXbtpI8tIg/pZw0jhuIHEpNG8XXAatv4=; z_c0=2|1:0|10:1752831125|4:z_c0|92:Mi4xZF9jckd3QUFBQUFBSUZKblhkX0FHU1lBQUFCZ0FsVk5rMkZuYVFDTlhJTWFpejRVV21rZUxvbFg1UGdVUzF0UXZ3|86d74e0522b0f6da6a5961fabcdd43c0d2d6d44cd016a4c04dca7526ab4d2029; Hm_lvt_98beee57fd2ef70ccdd5ca52b9740c49=1752562705,1752830746,1752831978,1752834129; HMACCOUNT=6690A0DA60280782; q_c1=db64099d329349adb445598247e443d1|1752834350000|1752834350000; SESSIONID=IDjlWFNb1UNTHiZeHIbAtq3HPpgsaBnLj5Q8MbD2yYM; JOID=V10TAEg3UhWRlZeEdjGkQXdAOJBtfhV6_cTRyREAB1jA3fjeEBt1av-TlIJz1Mt5_TcWvGi445H9zHGb0fzgIQI=; osd=UVERAE8xXheRkpGIdDGjR3tCOJdrchd6-sLdyxEHAVTC3f_YHBl1bfmfloJ00sd7_TAQsGq45JfxznGc1_DiIQU=; Hm_lpvt_98beee57fd2ef70ccdd5ca52b9740c49=1752835725; BEC=6bca8f185b99e85d761c7a0d8d692864; tst=h',
}

params = {
    'limit': '50',
    'desktop': 'true',
}

def save_xlsx(ranks,filename="zhihu_hotlist.xlsx"):#覆盖形式写入xlsx
    df = pd.DataFrame(ranks)
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    scientific_format = "0.00E+00"

    # 找到 "热度值" 列的索引（假设是第 3 列，即 'C' 列）
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):  # 从第 2 行开始（跳过表头）
        for cell in row:
            cell.number_format = scientific_format

    wb.save(filename)
    print(f"数据已保存到 {filename}")


def get_data(json_content):#清洗数据
    try:
        print("开始清洗数据")
        ranks=[]
        rank_list=json_content["data"]
        for one_rank in rank_list:
            title=one_rank["target"]["title_area"]["text"]
            link=one_rank["target"]["link"]["url"]
            hot_num=one_rank["target"]["metrics_area"]["text"]
            parts = hot_num.split()
            num_part = parts[0]
            if "万" in hot_num:
                num = int(num_part.replace('万', '')) * 10000
            else:
                num = int(num_part)
            print("标题:",title,"  链接:",link,"  热度值:",num)
            ranks.append({
                "标题":title,
                "链接":link,
                "热度值":num
            })
        return ranks
    except Exception as e:
        print("清洗数据错误")
        return []


def get_content():#获取html内容
    try:
        print("获取内容")
        response = requests.get(
            'https://www.zhihu.com/api/v3/feed/topstory/hot-lists/total',
            params=params,
            cookies=cookies,
            headers=headers,
        )
        json_content=response.json()
        return json_content
    except Exception as e:
        print("获取内容错误")
        return []

def main():#主函数
    print("函数执行中....")
    json_content=get_content()
    ranks=get_data(json_content)
    save_xlsx(ranks)
    print("一小时后继续执行")

if __name__=="__main__":
    main()
    schedule.every().hour.do(main)
    while True:
        schedule.run_pending()  # 检查是否有任务要执行
        time.sleep(60)