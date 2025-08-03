import time
import math
import random
import codecs
import base64
import asyncio
import aiohttp
import json
from Crypto.Cipher import AES
from datetime import datetime
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows

class Usercomment:
    def __init__(self, name, content, like, content_time):
        self.name = name
        self.content = content
        self.like = like
        self.content_time = content_time

    def to_dict(self):
        return {
            "用户名": self.name,
            "评论内容": self.content,
            "点赞数": self.like,
            "评论时间": self.content_time
        }


# 生成16个随机字符
def generate_random_strs(length):
    string = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    random_strs = ""
    for _ in range(length):
        e = random.random() * len(string)
        e = math.floor(e)
        random_strs += string[e]
    return random_strs

# AES加密
def AESencrypt(msg, key):
    padding = 16 - len(msg) % 16
    msg = msg + padding * chr(padding)
    iv = '0102030405060708'
    if isinstance(key, str):
        key = key.encode('utf-8')
    if isinstance(iv, str):
        iv = iv.encode('utf-8')
    cipher = AES.new(key, AES.MODE_CBC, iv)
    encryptedbytes = cipher.encrypt(msg.encode('utf-8'))
    encodestrs = base64.b64encode(encryptedbytes)
    enctext = encodestrs.decode('utf-8')
    return enctext

# RSA加密
def RSAencrypt(randomstrs, key, f):
    string = randomstrs[::-1]
    text = bytes(string, 'utf-8')
    seckey = int(codecs.encode(text, encoding='hex'), 16) ** int(key, 16) % int(f, 16)
    return format(seckey, 'x').zfill(256)

# 获取加密参数
def get_params(page):
    offset = (page - 1) * 20
    msg = '{"offset":' + str(offset) + ',"total":"True","limit":"20","csrf_token":""}'
    key = '0CoJUm6Qyw8W8jud'
    f = '00e0b509f6259df8642dbc35662901477df22677ec152b5ff68ace615bb7b725152b3ab17a876aea8a5aa76d2e417629ec4ee341f56135fccf695280104e0312ecbda92557c93870114af6c9d05c4f7f0c3685b7a46bee255932575cce10b424d813cfe4875d3e82047b97ddef52741d546b8e289dc6935b3ece0462db0a22b8e7'
    e = '010001'
    enctext = AESencrypt(msg, key)
    i = generate_random_strs(16)
    encText = AESencrypt(enctext, i)
    encSecKey = RSAencrypt(i, e, f)
    return encText, encSecKey


async def wait_time(min_sec=2, max_sec=4):#等待时间，防止反爬
    await asyncio.sleep(random.uniform(min_sec, max_sec))

def parse_json(html_json, all_content):#提取数据
    try:
        if not html_json or "comments" not in html_json:
            raise ValueError("无效的响应数据：缺少评论信息")

        for one_comment in html_json["comments"]:
            name = one_comment["user"]["nickname"]
            content = one_comment["content"]
            like = one_comment["likedCount"]
            timestamp = one_comment["time"]

            seconds = timestamp / 1000
            date_time = datetime.fromtimestamp(seconds)
            content_time = date_time.strftime("%Y-%m-%d %H:%M:%S")

            o_comment = Usercomment(name, content, like, content_time)
            all_content.append(o_comment.to_dict())
    except Exception as e:
        raise RuntimeError(f"解析数据失败: {str(e)}") from e


HEADERS = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded',
    'Host': 'music.163.com',
    'Origin': 'https://music.163.com',
    'Referer': 'https://music.163.com/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'Cookie': 'NMTID=00OYsn789EpzSpdsUlsk3cM87bhBeAAAAGYHfl8Qw; WEVNSM=1.0.0; WNMCID=ioanlx.1752849547001.01.0; sDeviceId=YD-%2FOpb%2B7YSRoREBwABRFeHgUKmupLU10J8; ntes_utid=tid._.ueRvZQTmXkpEQwAFUUKDhFKmvtaF1Mwa._.0; WM_TID=H052BJETwC1BQRBUFRPGgBaiu8eFegZ6; __snaker__id=AT3AHcqBltPHhwdn; JSESSIONID-WYYY=R%2F9UWBUZoOR4W0pdWK3zxDdYzF2XQEMIjIRfmJvRRC%2B234Pqe94J4YokE0w2%5Ctv43tGcSSKboBd5TIYsvWnfmoRxGJKegip9G5f5YdyU3TRkXYbyCUOWP%2FXHMVqX9KN5PketrGo8pwOxP6ATAbWK1K6EjBM6P5P0UYVkNS4bSct%2FA5M7%3A1752990382843; MUSIC_U=00A92DFDEF2591387F37DA4D642BCF90A3419A15C31974F3283426FC2505063CC0D8EF37D1FF60DD40463224572C56F30EDBC1F270F798BBCEA0F25FF5231305008EAAC99FD903A7DCD8CAF8BCAB4EC9C988837F7DE39D5C36DD8C854691682CE87876F3F32FF071C7F92366CE108030669E4D9041159277AD8C5374D7614AEF6496532DCCED168431C75AFD58CC90D37D16CDA396A9B898B585032DA3635C552F565B7E19D81BB5FB49613941BF0AFE283AB3EDA688D05F76EA379F60E82D87CDB5361D026A2763C2A258653E58A2C68B322DF490D52973F175B026281A606411F305745BAE7CD0DEE4DC898DFD27A96C0EF8B102EB98EFA3D6296ECBEEA5F0BBAD104ECC60BE86D08FC19F4292F3DEA9C658DD6274AA85694121E9841350DE4A196F88958EA3D685F38865A6EFC7247DDC4747FFB50BC5FDCB131011CA4827AD763165C330B6E54A40093779D130C86E3CA311FDD835F30CC54DEBF8B4B5166BDD59BDE943C9383012018737E726AB455D78BF18725D6C52B355DA130EBC4DC55E154BB925B337395814FEDDB85A8DC1; __remember_me=true; __csrf=cb8b5b8727af187fc90e94e85e944d39'
}


async def fetch_page(session, songid, page, retry=3):
    for attempt in range(retry):
        try:
            print(f"开始请求歌曲 {songid} 第 {page} 页 (尝试 {attempt + 1}/{retry})")
            params, encSecKey = get_params(page)
            url = f'https://music.163.com/weapi/v1/resource/comments/R_SO_4_{songid}?csrf_token='
            data = {'params': params, 'encSecKey': encSecKey}
            # 使用更长的超时时间
            timeout = aiohttp.ClientTimeout(total=30)
            async with session.post(url, data=data, headers=HEADERS, timeout=timeout) as response:
                # 先尝试以文本形式读取响应
                response_text = await response.text()
                # 检查响应是否包含反爬验证内容
                if "验证" in response_text or "security" in response_text.lower():
                    raise RuntimeError("触发反爬验证机制")
                # 尝试解析为JSON
                try:
                    response_json = json.loads(response_text)
                    if "code" in response_json and response_json["code"] != 200:
                        raise RuntimeError(f"API返回错误: {response_json.get('message', '未知错误')}")
                    return page, response_json
                except json.JSONDecodeError:
                    # 如果不是JSON，记录响应内容的前200字符用于调试
                    sample = response_text[:200] + "..." if len(response_text) > 200 else response_text
                    raise RuntimeError(f"响应不是有效的JSON: {sample}")

        except Exception as e:
            if attempt < retry - 1:
                # 指数退避策略
                wait = min(10, 2 ** (attempt + 1))
                print(f"请求失败，{wait}秒后重试: {str(e)}")
                await asyncio.sleep(wait + random.random())
                continue
            else:
                raise RuntimeError(f"歌曲 {songid} 第 {page} 页请求失败: {str(e)}") from e


async def get_content(songid):
    all_content = []
    print(f"\n开始爬取歌曲 {songid} 的评论")
    async with aiohttp.ClientSession() as session:
        # 创建5个并发任务
        tasks = [fetch_page(session, songid, p) for p in range(1, 6)]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        # 处理结果
        for result in results:
            if isinstance(result, Exception):
                print(f"错误: {str(result)}")
                continue
            page, html_json = result
            try:
                print(f"解析歌曲 {songid} 第 {page} 页数据")
                parse_json(html_json, all_content)
            except Exception as e:
                print(f"解析歌曲 {songid} 第 {page} 页数据失败: {str(e)}")
    print(f"完成歌曲 {songid} 的爬取，共获取 {len(all_content)} 条评论")
    return all_content


def save_excel(all_content, songid, xlsx_name):
    if not all_content:
        print(f"歌曲 {songid} 没有评论数据，跳过保存")
        return
    df = pd.DataFrame(all_content)
    try:
        if os.path.exists(xlsx_name):
            # 使用openpyxl直接操作Excel文件
            from openpyxl import load_workbook
            book = load_workbook(xlsx_name)
            # 删除已存在的同名工作表
            if str(songid) in book.sheetnames:
                del book[str(songid)]
            # 创建新工作表
            ws = book.create_sheet(str(songid))
            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            book.save(xlsx_name)
        else:
            df.to_excel(xlsx_name, sheet_name=str(songid), index=False)
        print(f"成功保存歌曲 {songid} 的 {len(all_content)} 条评论到 {xlsx_name}")
    except Exception as e:
        raise RuntimeError(f"保存歌曲 {songid} 数据到Excel失败: {str(e)}") from e

async def process_song(songid, xlsx_name):
    #获取数据和保存到xlsx里
    try:
        all_content = await get_content(songid)
        save_excel(all_content, songid, xlsx_name)
    except Exception as e:
        print(f"处理歌曲 {songid} 时发生严重错误: {str(e)}")
    finally:
        # 每首歌之间等待更长的时间
        await wait_time(4, 6)

async def main():
    xlsx_name = '网易云评论.xlsx'
    music_id = input("请输入音乐id，用空格隔开: ")
    music_id_list = music_id.split()
    print(f"开始爬取 {len(music_id_list)} 首歌曲的评论")
    # 顺序处理每首歌
    for index, songid in enumerate(music_id_list):
        try:
            print(f"\n处理歌曲 {songid} ({index + 1}/{len(music_id_list)})")
            await process_song(songid, xlsx_name)
        except Exception as e:
            print(f"处理歌曲 {songid} 时发生未捕获的异常: {str(e)}")
            # 继续处理下一首
            continue
    print("\n所有歌曲处理完成！")

if __name__ == "__main__":
    # 设置更宽松的事件循环策略
    import platform
    if platform.system() == 'Windows':
        #强制使用 SelectorEventLoop 的策略
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(main())